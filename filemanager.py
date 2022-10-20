from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import json
import os


class FileManager:

    def __init__(self):
        self.wb = None
        try:
            os.remove("order.json")
        except FileNotFoundError:
            pass

    def export_to_xlsx(self, forward, reverse):
        """Responsible for exporting sequences to an excel file in a specific format."""
        try:
            self.wb = load_workbook("primers.xlsx")
        except FileNotFoundError:
            self.wb = Workbook()
            ws = self.wb.active
            ws.title = "homo sapiens"
            self.wb.create_sheet("mus musculus")
            self.wb.create_sheet("rattus norvegicus")
            for sheet in self.wb.worksheets:
                sheet['A1'] = "Gene"
                sheet['B1'] = "Sequence"
                sheet['C1'] = "Length"
                sheet['D1'] = "Label"
        except PermissionError:
            raise "Close the file"
        finally:
            ws = self.wb[forward.organism]
            ws = self._write_primer(ws, forward)
            ws = self._write_primer(ws, reverse)
            self._add_to_json({"sequence": forward.sequence, "label": forward.print_primer()})
            self._add_to_json({"sequence": reverse.sequence, "label": reverse.print_primer()})
            self.adjust_width()
            self.wb.save("primers.xlsx")

    @staticmethod
    def _add_to_json(dict_):
        """Helper method, that export primers to the 'order.json' file for possible order in the future."""
        try:
            with open("order.json", 'r') as order:
                data = json.load(order)
            data.append(dict_)
            with open("order.json", 'w') as order:
                json.dump(data, order, indent=2)
        except FileNotFoundError:
            with open("order.json", 'w') as file:
                obj = []
                obj.append(dict_)
                json.dump(obj, file, indent=2)

    @staticmethod
    def _write_primer(ws, primer):
        """Helper method for writing primers into the excel sheet."""
        r = ws.max_row + 1
        ws[f'A{r}'] = primer.name
        ws[f'B{r}'] = primer.sequence
        ws[f'C{r}'] = primer.length
        ws[f'D{r}'] = primer.print_primer()
        return ws

    def is_open(self):
        self.wb = load_workbook("primers.xlsx")

    def adjust_width(self):
        """Adjusts width of columns in the excel sheet"""
        for sheet in self.wb.worksheets:
            dim_holder = DimensionHolder(worksheet=sheet)
            for col in range(sheet.min_column, sheet.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=40)
            sheet.column_dimensions = dim_holder






